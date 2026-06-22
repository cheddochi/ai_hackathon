"""
Excel Profit Sheet 파서
openpyxl 기반으로 .xlsx 파일에서 매출/매입 항목 추출
"""
from dataclasses import dataclass, field
from typing import List, Optional
import openpyxl
from app.services.pdf_parser import ParsedCharge, ParsedProfitSheet


# PROFIT/LOSS SHEET 공통 차지 코드 → (영문명, 한글명)
KNOWN_CHARGES: dict = {
    "OF":   ("OCEAN FREIGHT", "해상운임"),
    "BF":   ("BAF", "유류할증료"),
    "BAF":  ("BUNKER ADJUSTMENT FACTOR", "유류할증료"),
    "WAF":  ("WAR RISK SURCHARGE", "전쟁위험할증료"),
    "THC":  ("TERMINAL HANDLING CHARGE", "터미널처리비"),
    "CIC":  ("CONTAINER IMBALANCE CHARGE", "컨테이너불균형할증료"),
    "EBS":  ("EMERGENCY BUNKER SURCHARGE", "긴급유류할증료"),
    "CRS":  ("CURRENCY RISK SURCHARGE", "환위험할증료"),
    "EFS":  ("EMERGENCY FUEL SURCHARGE", "긴급연료할증료"),
    "DOC":  ("DOCUMENTATION FEE", "서류비"),
    "BL":   ("B/L FEE", "선하증권발행비"),
    "DO":   ("DELIVERY ORDER FEE", "화물인도지시서비"),
    "SEAL": ("SEAL FEE", "봉인비"),
    "AFR":  ("ADVANCE FILING RULES", "사전신고비"),
    "AMS":  ("AUTOMATED MANIFEST SYSTEM", "자동화물확인"),
    "ENS":  ("ENTRY SUMMARY DECLARATION", "수입신고"),
    "ISPS": ("INTERNATIONAL SHIP & PORT FACILITY SECURITY", "보안할증료"),
    "DUTY": ("CUSTOMS DUTY", "관세"),
    "VAT":  ("VALUE ADDED TAX", "부가세"),
    "AF":   ("AIR FREIGHT", "항공운임"),
    "FSC":  ("FUEL SURCHARGE", "연료할증료"),
    "SSC":  ("SECURITY SURCHARGE", "보안할증료"),
    "AWB":  ("AIR WAYBILL FEE", "항공운송장비"),
    "CC":   ("CUSTOMS CLEARANCE", "통관비"),
    "DL":   ("DELIVERY", "배달비"),
    "ST":   ("STORAGE", "창고비"),
    "HN":   ("HANDLING", "핸들링비"),
}


HEADER_ALIASES = {
    "charge_code": ["코드", "CODE", "항목코드", "CHARGE CODE", "ITEM"],
    "charge_name": ["항목명", "NAME", "DESCRIPTION", "DESC", "명칭"],
    "revenue": ["매출", "REVENUE", "SELL", "売上", "CHARGE"],
    "cost": ["매입", "COST", "BUY", "仕入", "PURCHASE"],
    "currency": ["통화", "CURRENCY", "CCY"],
    "partner": ["파트너", "PARTNER", "VENDOR", "업체"],
    "unit": ["단위", "UNIT", "QTY"],
}


def parse_excel(file_path: str) -> ParsedProfitSheet:
    """Excel 파일을 파싱하여 ParsedProfitSheet 반환"""
    result = ParsedProfitSheet()

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        result.parse_warnings.append(f"Excel 파일 열기 실패: {str(e)}")
        return result

    # 첫 번째 시트 또는 'Profit', 'P&L' 이름 시트 우선
    target_sheet = None
    for name in wb.sheetnames:
        if any(k in name.upper() for k in ["PROFIT", "P&L", "PL", "손익"]):
            target_sheet = wb[name]
            break
    if target_sheet is None:
        target_sheet = wb.active

    rows = list(target_sheet.iter_rows(values_only=True))
    if not rows:
        result.parse_warnings.append("시트에 데이터가 없습니다.")
        return result

    # 메타 정보 추출 (상단 영역 스캔)
    result = _extract_meta_from_excel(rows, result)

    # 헤더 행 찾기
    header_row_idx, col_map = _find_header_row(rows)

    if header_row_idx < 0 or not col_map:
        # 헤더 없으면 텍스트 기반 추출 시도
        result.charges = _fallback_parse(rows)
        result.parse_warnings.append("헤더 자동 감지 실패 — 패턴 기반 파싱 사용")
        return result

    # 데이터 행 파싱
    charges = []
    for row in rows[header_row_idx + 1:]:
        if not row or not any(row):
            continue
        charge = _parse_data_row(row, col_map)
        if charge:
            charges.append(charge)

    result.charges = charges
    return result


def _extract_meta_from_excel(rows: list, result: ParsedProfitSheet) -> ParsedProfitSheet:
    """상단 영역에서 메타 정보 추출"""
    for row in rows[:20]:  # 상단 20행만 스캔
        for i, cell in enumerate(row):
            if cell is None:
                continue
            cell_str = str(cell).strip()
            next_val = str(row[i + 1]).strip() if i + 1 < len(row) and row[i + 1] else ""

            cell_upper = cell_str.upper()
            if any(k in cell_upper for k in ["CASE", "案件", "안건", "H.B/L", "HBL"]) and next_val:
                result.hbl_no = next_val
            elif any(k in cell_upper for k in ["CUSTOMER", "顧客", "거래처", "SHIPPER"]) and next_val:
                result.customer_name = next_val
            elif any(k in cell_upper for k in ["ASSIGNEE", "담당", "担当"]) and next_val:
                result.assignee_name = next_val
            elif any(k in cell_upper for k in ["FROM", "ORIGIN", "POL", "출발"]) and next_val:
                result.pol = next_val.upper()
            elif any(k in cell_upper for k in ["TO", "DEST", "POD", "도착"]) and next_val:
                result.pod = next_val.upper()
            elif any(k in cell_upper for k in ["WEIGHT", "G.W", "중량"]) and next_val:
                try:
                    result.weight_kg = float(str(next_val).replace(",", ""))
                except ValueError:
                    pass
            elif "CBM" in cell_upper and next_val:
                try:
                    result.cbm = float(str(next_val).replace(",", ""))
                except ValueError:
                    pass
    return result


def _find_header_row(rows: list):
    """헤더 행 및 컬럼 매핑 탐지"""
    for idx, row in enumerate(rows):
        if not row or not any(row):
            continue
        row_strs = [str(c or "").upper().strip() for c in row]
        col_map = {}

        has_revenue = False
        has_cost = False

        for col_idx, cell_str in enumerate(row_strs):
            for field_name, aliases in HEADER_ALIASES.items():
                if any(alias.upper() in cell_str for alias in aliases):
                    if field_name not in col_map:
                        col_map[field_name] = col_idx
                    if field_name == "revenue":
                        has_revenue = True
                    if field_name == "cost":
                        has_cost = True

        if has_revenue or has_cost:
            return idx, col_map

    return -1, {}


def _parse_data_row(row: tuple, col_map: dict) -> Optional[ParsedCharge]:
    """단일 데이터 행 파싱"""
    def get_cell(field_name):
        idx = col_map.get(field_name)
        if idx is not None and idx < len(row):
            return row[idx]
        return None

    code_raw = str(get_cell("charge_code") or "").strip().upper()
    name_raw = str(get_cell("charge_name") or "").strip()

    if not code_raw and not name_raw:
        return None

    charge_code = code_raw
    charge_name = name_raw
    if code_raw in KNOWN_CHARGES:
        charge_name = charge_name or KNOWN_CHARGES[code_raw][0]
    elif name_raw:
        for code, (en_name, _) in KNOWN_CHARGES.items():
            if en_name.upper() in name_raw.upper() or code in name_raw.upper():
                charge_code = code
                charge_name = charge_name or en_name
                break

    if not charge_code:
        return None

    rev_val = _to_float(get_cell("revenue"))
    cost_val = _to_float(get_cell("cost"))

    if rev_val and rev_val > 0:
        is_revenue = True
        amount = rev_val
    elif cost_val and cost_val > 0:
        is_revenue = False
        amount = cost_val
    else:
        return None

    currency = str(get_cell("currency") or "JPY").strip().upper() or "JPY"
    account = str(get_cell("partner") or "").strip()

    # amount_jpy: JPY면 그대로, 외화면 일단 amount로 (API에서 환율 적용)
    amount_jpy = amount if currency == "JPY" else 0.0

    return ParsedCharge(
        charge_code=charge_code,
        charge_name=charge_name,
        is_revenue=is_revenue,
        currency=currency,
        amount=amount,
        amount_jpy=amount_jpy,
        account_name=account,
        confidence=0.9,
    )


def _fallback_parse(rows: list) -> List[ParsedCharge]:
    """헤더 감지 실패 시 전체 셀 스캔"""
    charges = []
    for row in rows:
        if not row:
            continue
        row_text = " ".join(str(c or "") for c in row)
        for code, (name, _) in KNOWN_CHARGES.items():
            if code in row_text.upper():
                for cell in row:
                    try:
                        val = float(str(cell).replace(",", ""))
                        if val > 0:
                            charges.append(ParsedCharge(
                                charge_code=code,
                                charge_name=name,
                                is_revenue=True,
                                currency="JPY",
                                amount=val,
                                amount_jpy=val,
                                confidence=0.4,
                            ))
                            break
                    except (ValueError, TypeError):
                        continue
    return charges


def _to_float(val) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return None
